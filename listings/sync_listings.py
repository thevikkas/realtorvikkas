#!/usr/bin/env python3
"""Push the Excel 'Jaipur' sheet to the website.

Reads Realtor_Vikkas_Listings.xlsx → the Jaipur sheet → every row whose
'Listing Status' is 'Live', and POSTs them to the site's /api/sync-listings
endpoint. The site upserts them and hides everything else, so the website
ends up matching your control sheet.

Usage:
    python3 sync_listings.py                      # sync to local site (localhost:8000)
    SITE_URL=https://realtorvikkas.com python3 sync_listings.py   # sync to live
Env:
    SITE_URL  target site (default http://localhost:8000)
    SYNC_KEY  must match the site's SYNC_KEY (default matches the local fallback)
"""
import os
import json
import urllib.request
import urllib.parse
from openpyxl import load_workbook

EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Realtor_Vikkas_Listings.xlsx")
SITE = os.environ.get("SITE_URL", "http://localhost:8000").rstrip("/")
KEY = os.environ.get("SYNC_KEY", "vikkas-jaipur-8753")


def read_live_rows():
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["Jaipur"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[idx["Property Name"]] if row else None
        if not name:
            continue
        status = str(row[idx["Listing Status"]] or "").strip().lower()
        if status != "live":                 # only Live rows go to the website
            continue
        area = str(row[idx["Location/Area"]] or "")
        locality, _, city = area.partition(",")
        try:
            price = int(float(row[idx["Price"]] or 0))
        except (TypeError, ValueError):
            price = 0
        out.append({
            "title": str(name).strip(),
            "ptype": str(row[idx["Type"]] or "Flat").strip(),
            "city": (city.strip() or "Jaipur"),
            "locality": locality.strip(),
            "price": price,
            "photos_url": str(row[idx["Photos link"]] or "").strip(),
            "listing": "rent" if 0 < price < 200000 else "buy",
        })
    return out


def main():
    props = read_live_rows()
    print(f"→ syncing {len(props)} Live Jaipur listing(s) to {SITE} …")
    data = urllib.parse.urlencode({"key": KEY, "payload": json.dumps(props)}).encode()
    req = urllib.request.Request(SITE + "/api/sync-listings", data=data)
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            print("✓ site replied:", r.read().decode())
    except urllib.error.HTTPError as ex:
        print(f"✗ HTTP {ex.code}:", ex.read().decode())
    except Exception as ex:  # noqa: BLE001
        print("✗ failed:", ex)


if __name__ == "__main__":
    main()
